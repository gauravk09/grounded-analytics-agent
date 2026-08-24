"""Make a mess on purpose, so the answer is known before ingest runs.

The whole point of evals-first here: we do NOT judge ingest by eyeballing its output, and we do
NOT compare it against a spec a human wrote (if both are wrong the same way, that passes). Instead
we start from a CLEAN table we can read by eye, deform it with code, and remember exactly what we
did. Ingest must undo the deformation and land every value back at the cell we placed it in.

Two deformations, because the two failing files are opposite errors:
    fold()      tidy -> cross-tab (PPAC-like). Ingest must MELT it back.
    keep()      tidy record table (climbing-like). Ingest must NOT melt it.

Each returns (xlsx_path, Truth). Truth is the answer key:
    rows      the clean rows as a SET of tuples — order must not matter
    cell_of   {(identifier..., variable): "B4"} — where each number was written

No model, no network. A ~5-row table clears propose()'s ">=3 numeric per column" floor while
staying small enough to check by hand.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


@dataclass
class Truth:
    rows: set[tuple]                 # clean rows, as a set — the row-set check
    cell_of: dict[tuple, str]        # (id..., variable) -> "Sheet cell" like "B4" — the receipt check
    should_melt: bool                # what the classifier must decide
    id_columns: list[str]            # which columns identify a row (the one question we'd ask)
    sheet: str = "S"


def fold(clean: list[dict], id_cols: list[str], var_col: str, val_col: str,
         out: Path, blank_rows: int = 0, footnote: str | None = None) -> tuple[Path, Truth]:
    """Turn a tidy table into a cross-tab: the values of `var_col` become the header, and
    `val_col` fills the matrix. This is the PPAC shape. Answer key = the original clean rows.

    blank_rows / footnote are knobs: a real government sheet has spacer rows and a source line,
    and moving numbers around must not change the answer. They make the answer key harder to hit
    by accident.
    """
    assert len(id_cols) == 1, "fold writes one identifier column (kept minimal on purpose)"
    id_col = id_cols[0]
    variables = list(dict.fromkeys(r[var_col] for r in clean))     # header order, de-duped
    entities = list(dict.fromkeys(r[id_col] for r in clean))
    lookup = {(r[id_col], r[var_col]): r[val_col] for r in clean}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"

    top = 1 + blank_rows
    ws.cell(top, 1, id_col)
    for j, v in enumerate(variables, start=2):
        ws.cell(top, j, v)

    cell_of: dict[tuple, str] = {}
    for i, ent in enumerate(entities, start=top + 1):
        ws.cell(i, 1, ent)
        for j, v in enumerate(variables, start=2):
            if (ent, v) in lookup:
                ws.cell(i, j, lookup[(ent, v)])
                cell_of[(ent, v)] = f"{get_column_letter(j)}{i}"
    if footnote:
        ws.cell(i + 2, 1, footnote)

    wb.save(out)
    rows = {(r[id_col], r[var_col], r[val_col]) for r in clean}
    return out, Truth(rows=rows, cell_of=cell_of, should_melt=True, id_columns=[id_col])


def keep(clean: list[dict], id_cols: list[str], measure_cols: list[str],
         out: Path) -> tuple[Path, Truth]:
    """Write a tidy RECORD table unchanged — the climbing shape. Answer key says: do NOT melt.
    One clean row per input row; each measure cell cited at its own address.
    """
    cols = id_cols + measure_cols
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    for j, name in enumerate(cols, start=1):
        ws.cell(1, j, name)

    cell_of: dict[tuple, str] = {}
    rows: set[tuple] = set()
    for i, r in enumerate(clean, start=2):
        key = tuple(r[c] for c in id_cols)
        for j, name in enumerate(cols, start=1):
            ws.cell(i, j, r[name])
        for name in measure_cols:
            cell_of[key + (name,)] = f"{get_column_letter(cols.index(name) + 1)}{i}"
        rows.add(tuple(r[c] for c in cols))
    wb.save(out)
    return out, Truth(rows=rows, cell_of=cell_of, should_melt=False, id_columns=id_cols)
