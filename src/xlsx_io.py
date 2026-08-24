"""One loader for every tabular file, so the rest of the pipeline never learns the format.

A CSV is read into an in-memory openpyxl workbook with numeric-looking cells coerced to numbers —
because the probe decides a column is a measure by seeing real numbers, not digit strings. After
this boundary, probe / propose / ingest / lineage are identical for .xlsx and .csv: a citation like
`Sheet1!B2` points at CSV row 2, column B, which is exactly where the value sits.

Deliberately conservative coercion: only a clean number becomes a number. "2008-09", an ISIN, a
date string — all stay text, because guessing them into numbers would corrupt both the classifier
and the citations.
"""
from __future__ import annotations

import csv as _csv
from pathlib import Path

import openpyxl
from openpyxl import Workbook

TABULAR = {".csv", ".tsv", ".txt"}


def _coerce(sval: str):
    s = sval.strip()
    if s == "":
        return None
    try:
        f = float(s)                       # plain numbers only; "1,234" and "2008-09" stay text
    except ValueError:
        return s
    return int(f) if f.is_integer() and "." not in s and "e" not in s.lower() else f


def _csv_workbook(path: Path) -> Workbook:
    delim = "\t" if path.suffix.lower() == ".tsv" else ","
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    with open(path, newline="", encoding="utf-8-sig") as f:
        for r, row in enumerate(_csv.reader(f, delimiter=delim), start=1):
            for c, val in enumerate(row, start=1):
                v = _coerce(val)
                if v is not None:
                    ws.cell(r, c, v)
    return wb


def load(path, data_only: bool = True, read_only: bool = False) -> Workbook:
    """Return an openpyxl workbook for a .xlsx OR a .csv/.tsv. CSVs have no formulas or merges, so
    data_only/read_only are accepted for signature parity and ignored."""
    p = Path(path)
    if p.suffix.lower() in TABULAR:
        return _csv_workbook(p)
    return openpyxl.load_workbook(p, data_only=data_only, read_only=read_only)
