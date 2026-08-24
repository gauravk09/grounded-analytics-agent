"""Read the PPAC workbook into DuckDB, writing a cell-level receipt book as we go.

The single rule this module exists to enforce: every value that lands in the tidy
table records the exact cell it came from, at the moment it is read. Lineage is a
fact about *how* the file was read, so it cannot be recovered afterwards.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import duckdb
import openpyxl
from openpyxl.utils import get_column_letter

# --- Layout of this specific workbook, found by inspection -------------------
HEADER_ROW = 8           # 'STATE/UT' | '2008-09' | ... | '2025-26'
FIRST_DATA_ROW = 9
LAST_DATA_ROW = 55       # row 56+ are footnotes inside the sheet body
NOTE_ROWS = [1, 5, 6, 7, 56, 57]   # title/units/source/footnotes — citable context, not data
LABEL_COL = 1            # column A
FIRST_YEAR_COL = 2       # column B
LAST_YEAR_COL = 19       # column S
TITLE_CELL = "A7"        # 'ALL PRODUCTS ANNUAL INDUSTRY SALES...' -> product provenance

SHEET_PRODUCTS = {
    "PT_Cons_Statewise": "ALL",
    "PT_Cons_Statewise MS": "MS",       # Motor Spirit = petrol
    "PT_Cons_Statewise HSD": "HSD",     # High Speed Diesel
}

FOOTNOTE_MARKER = re.compile(r"\[\d+\]|\*")


@dataclass
class Cell:
    """One receipt: where a single field of a single output row came from."""
    row_id: int
    column: str
    sheet: str
    a1: str
    raw_value: str | None
    formula: str | None


def clean_label(raw: str) -> str:
    """Strip footnote markers and stray whitespace. The raw form stays in the receipt."""
    return FOOTNOTE_MARKER.sub("", raw).strip()


def classify(label: str) -> str:
    """What kind of row is this? Aggregates are kept and tagged, never dropped."""
    if label.startswith("REGION - "):
        return "section_header"
    if label.lower() == "region total":
        return "region_total"
    if label.upper() == "ALL INDIA TOTAL":
        return "all_india"
    return "state"


def ingest(xlsx_path: Path, db_path: Path) -> None:
    # Two passes over the same file: openpyxl gives values OR formulas, never both.
    wb_values = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_formulas = openpyxl.load_workbook(xlsx_path, data_only=False)

    rows: list[dict] = []
    receipts: list[Cell] = []
    notes: list[tuple[str, str, str]] = []
    row_id = 0

    for sheet_name, product in SHEET_PRODUCTS.items():
        sv, sf = wb_values[sheet_name], wb_formulas[sheet_name]

        # Titles and footnotes are not data, but they explain the data. Keep them citable:
        # e.g. A57 is why the region totals do not sum to ALL INDIA TOTAL.
        for r in NOTE_ROWS:
            text = sv.cell(r, LABEL_COL).value
            if text is not None:
                notes.append((sheet_name, f"A{r}", str(text).strip()))

        # Year labels live in the header row, so each one has its own address.
        years = {
            col: (str(sv.cell(HEADER_ROW, col).value).strip(),
                  f"{get_column_letter(col)}{HEADER_ROW}")
            for col in range(FIRST_YEAR_COL, LAST_YEAR_COL + 1)
            if sv.cell(HEADER_ROW, col).value is not None
        }

        region = None          # forward-filled from the last section header
        region_a1 = None       # ...and we remember which cell it came from

        for r in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
            raw_label = sv.cell(r, LABEL_COL).value
            if raw_label is None:
                continue
            raw_label = str(raw_label)
            label = clean_label(raw_label)
            kind = classify(label)
            label_a1 = f"A{r}"

            if kind == "section_header":
                region = label.removeprefix("REGION - ").strip()
                region_a1 = label_a1
                continue  # carries no data of its own

            # Total rows are not states. Leaving their label in `state` would surface
            # "Region Total" as a selectable state name in the catalog.
            state = label if kind == "state" else None
            row_region = None if kind == "all_india" else region

            # The unpivot: one sheet row becomes one output row per year column.
            for col, (year, year_a1) in years.items():
                value = sv.cell(r, col).value
                if value is None:
                    continue
                formula = sf.cell(r, col).value
                value_a1 = f"{get_column_letter(col)}{r}"

                rows.append({
                    "__row_id": row_id,
                    "state": state,
                    "region": row_region,
                    "year": year,
                    "product": product,
                    "value": float(value),
                    "row_kind": kind,
                })

                receipts.append(Cell(row_id, "value", sheet_name, value_a1,
                                     str(value),
                                     formula if isinstance(formula, str) and formula.startswith("=") else None))
                receipts.append(Cell(row_id, "year", sheet_name, year_a1, year, None))
                receipts.append(Cell(row_id, "product", sheet_name, TITLE_CELL,
                                     str(sv[TITLE_CELL].value), None))
                # row_kind was read off the label cell, so it cites that cell even when
                # the row has no state of its own.
                receipts.append(Cell(row_id, "row_kind", sheet_name, label_a1, raw_label, None))
                if state is not None:
                    receipts.append(Cell(row_id, "state", sheet_name, label_a1, raw_label, None))
                if row_region is not None:
                    # Note the address: the region came from the section header row, not this one.
                    receipts.append(Cell(row_id, "region", sheet_name, region_a1, region, None))

                row_id += 1

    _write(db_path, rows, receipts, notes)
    print(f"{len(rows)} rows, {len(receipts)} receipts, {len(notes)} notes -> {db_path}")


def _write(db_path: Path, rows: list[dict], receipts: list[Cell],
           notes: list[tuple[str, str, str]]) -> None:
    con = duckdb.connect(str(db_path))
    con.execute("DROP TABLE IF EXISTS consumption")
    con.execute("DROP TABLE IF EXISTS cell_map")
    con.execute("DROP TABLE IF EXISTS sheet_notes")
    con.execute("""
        CREATE TABLE consumption (
            __row_id INTEGER, state VARCHAR, region VARCHAR,
            year VARCHAR, product VARCHAR, value DOUBLE, row_kind VARCHAR
        )""")
    con.execute("""
        CREATE TABLE cell_map (
            table_name VARCHAR, row_id INTEGER, column_name VARCHAR,
            sheet VARCHAR, a1 VARCHAR, raw_value VARCHAR, formula VARCHAR
        )""")
    con.execute("CREATE TABLE sheet_notes (sheet VARCHAR, a1 VARCHAR, text VARCHAR)")
    con.executemany("INSERT INTO sheet_notes VALUES (?,?,?)", notes)
    con.executemany(
        "INSERT INTO consumption VALUES (?,?,?,?,?,?,?)",
        [tuple(r.values()) for r in rows],
    )
    con.executemany(
        "INSERT INTO cell_map VALUES ('consumption',?,?,?,?,?,?)",
        [(c.row_id, c.column, c.sheet, c.a1, c.raw_value, c.formula) for c in receipts],
    )
    con.close()


if __name__ == "__main__":
    root = Path(__file__).resolve().parent.parent
    ingest(root / "data" / "ppac_statewise_sales.xlsx", root / "data" / "warehouse.duckdb")
