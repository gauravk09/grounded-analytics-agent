"""The universal half of ingest: read any sheet's STRUCTURE, interpret nothing.

This is the part that works on a workbook nobody has seen, because it makes no claims. It
reports what is there — how many cells a row fills, whether they are text or numbers, whether
the row is merged, whether its label looks like a total — and leaves every judgement to the
proposer and the human.

The split matters: guessing structure silently is hallucination relocated into the ingest layer,
where a wrong header row makes every number *and* every citation wrong, confidently. So the
reading is universal and the interpreting is confirmed.

No values are reported. A profile carries counts, types and labels — never a figure — for the
same reason the catalog does not (D17).
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field

import openpyxl  # noqa: F401  (kept for type parity)
from openpyxl.utils import get_column_letter
from xlsx_io import load as _load

TOTAL_WORDS = re.compile(r"\b(total|totals|sum|subtotal|grand|all india|overall)\b", re.I)
SECTION_HINT = re.compile(r"^\s*(region|zone|area|group|category|section)\b[\s:\-–]", re.I)


@dataclass
class RowProfile:
    row: int
    filled: int                  # cells with something in them
    numeric: int                 # of those, how many are numbers
    label: str | None            # the leftmost text cell, cleaned
    merged: bool
    looks_total: bool            # label contains total/sum/subtotal
    looks_section: bool          # label looks like a section heading
    # Text cells by column. The ENTITY label is not always the leftmost one — the Union Budget
    # keeps numbering in B and names in C/D — so callers must be able to read the column the
    # geometry step actually chose.
    texts: dict = field(default_factory=dict)

    @property
    def text(self) -> int:
        return self.filled - self.numeric


@dataclass
class SheetProfile:
    name: str
    max_row: int
    max_col: int
    merged: int
    formulas: int
    rows: list[RowProfile] = field(default_factory=list)
    col_text: dict[int, int] = field(default_factory=dict)     # column -> text cell count
    col_numeric: dict[int, int] = field(default_factory=dict)

    def _line(self, r: RowProfile, cols: int) -> str:
        cells = "".join(r.cells[:cols]) if hasattr(r, "cells") else ""
        return f"{r.row:>4} {cells:<{cols}} {'M' if r.merged else ' '} {(r.label or '')[:46]}"

    def grid(self, rows: int = 30, cols: int = 12) -> str:
        """The first `rows` rows, verbatim. For a HUMAN — the confirmation screen shows the sheet
        as it is, because a reviewer checking our geometry needs to see what we did not keep.

        Shows SHAPE, not data: `T` a text cell, `#` a number, `f` a formula, `.` empty — plus
        the leftmost label, which is the one piece of content needed to recognise a header or
        a subtotal.
        """
        return "\n".join(self._line(r, cols) for r in self.rows[:rows])

    def anchors(self, edge: int = 4) -> set[int]:
        """Row numbers worth showing a model: the boundaries, not the bulk.

        A row earns its place two independent ways, and both are needed:

        **Its shape changed.** Fifty consecutive `T#########` rows say exactly what the first one
        said. What carries information is where a run *starts or stops* — data begins, the header
        sits above it, a section breaks, the table ends. We keep the row either side of each change
        so the last row of the old run is visible too — that is what locates `last_data_row`.
        Formula cells are normalised to numbers first: `=SUM(...)` and `4996.7` are the same
        structural fact.

        **Its label is unusual.** Shape alone is blind to the single most important row type: a
        subtotal has the *same* shape as the data it sums (`T#########`), so change-detection drops
        it silently. Merged, total-like and section-like rows are kept on content instead.

        Plus the first and last few rows unconditionally — titles live at the top, footnotes and
        grand totals at the bottom, and "row 500 is the last row" versus "row 500 is a total" is a
        distinction you can only make by seeing its neighbours.
        """
        keep: set[int] = set()
        rows = self.rows
        if not rows:
            return keep

        for r in rows[:edge] + rows[-edge:]:
            keep.add(r.row)

        prev = None
        for i, r in enumerate(rows):
            # `f` and `#` are both "a number lives here". Whether it was typed or computed is a
            # fact about the cell, not about where the table starts — treating it as a boundary
            # made every stray formula cost three rows.
            shape = getattr(r, "cells", "").replace("f", "#")
            if prev is not None and shape != prev:
                # The changed row, and the last row of the run before it — that pair is what
                # locates an edge. The row AFTER the change adds nothing the changed row didn't.
                for j in (i - 1, i):
                    keep.add(rows[j].row)
            prev = shape
            if r.merged or r.looks_total or r.looks_section:
                keep.add(r.row)
        return keep

    def sketch(self, cols: int = 12, edge: int = 4) -> str:
        """`anchors()` rendered, with each dropped run replaced by a count.

        The count matters: "40 more rows, same shape" tells the model forty data rows were there
        without spending forty lines saying so. Cost scales with how COMPLICATED a sheet is, not
        how big — a 5,000-row sheet and a 50-row sheet with the same structure cost the same.
        """
        keep = self.anchors(edge)
        out: list[str] = []
        run = 0
        for r in self.rows:
            if r.row in keep:
                if run:
                    out.append(f"     … {run} more rows, same shape")
                    run = 0
                out.append(self._line(r, cols))
            else:
                run += 1
        if run:
            out.append(f"     … {run} more rows, same shape")
        return "\n".join(out)


def profile(path, sheet_name: str, scan_rows: int = 200) -> SheetProfile:
    wb_v = _load(path, data_only=True)
    wb_f = _load(path, data_only=False)
    sv, sf = wb_v[sheet_name], wb_f[sheet_name]

    merged_rows = {r for rng in sf.merged_cells.ranges for r in range(rng.min_row, rng.max_row + 1)}
    p = SheetProfile(name=sheet_name, max_row=sv.max_row, max_col=sv.max_column,
                     merged=len(sf.merged_cells.ranges), formulas=0)

    for r in range(1, min(sv.max_row, scan_rows) + 1):
        filled = numeric = 0
        label = None
        marks = []
        texts: dict[int, str] = {}
        for c in range(1, min(sv.max_column, 40) + 1):
            v = sv.cell(r, c).value
            f = sf.cell(r, c).value
            is_formula = isinstance(f, str) and f.startswith("=")
            p.formulas += is_formula
            if v is None or (isinstance(v, str) and not v.strip()):
                marks.append(".")
                continue
            filled += 1
            if isinstance(v, (int, float)):
                numeric += 1
                marks.append("f" if is_formula else "#")
                p.col_numeric[c] = p.col_numeric.get(c, 0) + 1
            else:
                marks.append("T")
                p.col_text[c] = p.col_text.get(c, 0) + 1
                texts[c] = str(v).strip()
                if label is None:
                    label = str(v).strip()

        rp = RowProfile(row=r, filled=filled, numeric=numeric, label=label,
                        merged=r in merged_rows,
                        looks_total=bool(label and TOTAL_WORDS.search(label)),
                        looks_section=bool(label and SECTION_HINT.match(label)),
                        texts=texts)
        rp.cells = "".join(marks)      # attached for grid(); not part of the dataclass contract
        p.rows.append(rp)
    return p


def sheets(path) -> list[str]:
    return _load(path, read_only=True).sheetnames


def column_letters(profile: SheetProfile) -> dict[str, list[str]]:
    """Which columns are mostly text and which mostly numeric, as letters."""
    text = [get_column_letter(c) for c, n in sorted(profile.col_text.items())
            if n > profile.col_numeric.get(c, 0)]
    num = [get_column_letter(c) for c, n in sorted(profile.col_numeric.items())
           if n >= profile.col_text.get(c, 0)]
    return {"text": text, "numeric": num}
