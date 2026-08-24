"""Generic ingest: read ANY workbook, given a confirmed spec.

Same two-pass read and same cell-level receipts as before (D6, D7) — the difference is that the
constants now come from `specs/<file>.yaml` instead of being typed into the source. That is the
entire Phase 1 change, and it is what turns "works on the file I read by hand" into "works on a
file nobody has seen, once someone confirms what it means".

The invariant this file must not break: every value written to the tidy table records its source
cell in the SAME loop iteration that read it. Lineage is a fact about how the file was read, and
there is no later moment at which it can be recovered.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import duckdb
import openpyxl  # noqa: F401
from openpyxl.utils import column_index_from_string, get_column_letter

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent))
from spec import SheetSpec, Spec                                  # noqa: E402
from xlsx_io import load as _load                                # noqa: E402

FOOTNOTE_MARKER = re.compile(r"\[\d+\]|\*")


@dataclass
class Cell:
    row_id: int
    column: str
    sheet: str
    a1: str
    raw_value: str | None
    formula: str | None


def clean_label(raw: str) -> str:
    """Strip footnote markers and stray whitespace. The raw form stays in the receipt (D8)."""
    return FOOTNOTE_MARKER.sub("", raw).strip()


def classify(label: str, sheet: SheetSpec) -> str:
    """Which kind of row is this? Aggregates are kept and tagged, never dropped (D9)."""
    for kind, pattern in sheet.row_kinds.items():
        if re.search(pattern, label, re.I):
            return kind
    return "entity"


def ingest(spec: Spec, xlsx_path: Path, db_path: Path) -> dict:
    # Two passes over the same file: openpyxl gives values OR formulas, never both (D7).
    wb_v = _load(xlsx_path, data_only=True)
    wb_f = _load(xlsx_path, data_only=False)

    layouts = {s.layout for s in spec.sheets}
    if len(layouts) > 1:
        # A union schema across mixed layouts is out of scope for now. Raise rather than emit a
        # half-melted table that looks fine and is wrong (principle: an interface that can be
        # misused quietly will be).
        raise ValueError(f"workbook mixes layouts {layouts}; ingest one layout per workbook")
    if "record" in layouts:
        return _ingest_record(spec, wb_v, wb_f, db_path)

    dims = [spec.entity]
    if any(s.section_dimension for s in spec.sheets):
        dims.append(next(s.section_dimension for s in spec.sheets if s.section_dimension))
    dims.append(spec.period)
    for s in spec.sheets:
        for k in s.constants:
            if k not in dims:
                dims.append(k)
    dims.append(spec.measure)
    dims.append("row_kind")

    rows: list[dict] = []
    receipts: list[Cell] = []
    notes: list[tuple[str, str, str]] = []
    row_id = 0

    for sheet in spec.sheets:
        sv, sf = wb_v[sheet.name], wb_f[sheet.name]
        label_col = column_index_from_string(sheet.label_column)
        lo = column_index_from_string(sheet.first_value_column)
        hi = column_index_from_string(sheet.last_value_column)

        # Titles, units, sources — not data, but they explain the data, so keep them citable.
        for r in sheet.note_rows:
            for c in range(1, lo):
                text = sv.cell(r, c).value
                if text is not None and str(text).strip():
                    notes.append((sheet.name, f"{get_column_letter(c)}{r}", str(text).strip()))
                    break

        # Period labels live in the header row(s). A stacked header (year over estimate-type)
        # combines top-to-bottom into ONE unique label — otherwise two columns both read
        # "Budget Estimates" and collide. Cited at the topmost cell, the one the single-row
        # read used to miss. Single-row files have header_rows empty, so this is unchanged.
        hrows = sheet.header_rows or [sheet.header_row]

        def _period(col):
            parts = []
            for hr in hrows:
                v = sv.cell(hr, col).value
                if v is not None and str(v).strip():
                    parts.append(" ".join(str(v).split()))   # collapse newlines/whitespace
            return " ".join(parts), f"{get_column_letter(col)}{hrows[0]}"

        periods = {
            col: _period(col)
            for col in range(lo, hi + 1)
            if any(sv.cell(hr, col).value is not None for hr in hrows)
        }

        section = section_a1 = None

        for r in range(sheet.first_data_row, sheet.last_data_row + 1):
            raw_label = sv.cell(r, label_col).value
            if raw_label is None or not str(raw_label).strip():
                continue
            raw_label = str(raw_label)
            label = clean_label(raw_label)
            kind = classify(label, sheet)
            label_a1 = f"{sheet.label_column}{r}"

            if kind == "section_header":
                # A section defines a dimension for the rows beneath it, and its CELL is what
                # those rows cite for that dimension (D10).
                section = re.sub(sheet.row_kinds["section_header"], "", label, flags=re.I).strip()
                section_a1 = label_a1
                continue

            is_entity = kind == "entity"
            entity = label if is_entity else None
            sec = section if kind != "grand_total" else None

            # The unpivot: one sheet row becomes one output row per period column. Each output
            # row's address is generated here, because `col` and `r` exist nowhere else (D6).
            for col, (period, period_a1) in periods.items():
                value = sv.cell(r, col).value
                if not isinstance(value, (int, float)):
                    continue                      # '..' and other missing-value markers
                formula = sf.cell(r, col).value
                value_a1 = f"{get_column_letter(col)}{r}"

                rec: dict = {"__row_id": row_id}
                if entity is not None:
                    rec[spec.entity] = entity
                else:
                    rec[spec.entity] = None
                if sheet.section_dimension:
                    rec[sheet.section_dimension] = sec
                rec[spec.period] = period
                rec.update(sheet.constants)
                rec[spec.measure] = float(value)
                # "entity", never spec.entity. A row-kind vocabulary that borrows a word from
                # the data creates collisions: when row_kind had a value "state" AND a column was
                # called "state", "which state used the most" was ambiguous between them (D34).
                # Generic kinds make that class of bug impossible rather than handled.
                rec["row_kind"] = "entity" if is_entity else kind
                rows.append(rec)

                receipts.append(Cell(row_id, spec.measure, sheet.name, value_a1, str(value),
                                     formula if isinstance(formula, str)
                                     and formula.startswith("=") else None))
                receipts.append(Cell(row_id, spec.period, sheet.name, period_a1, period, None))
                receipts.append(Cell(row_id, "row_kind", sheet.name, label_a1, raw_label, None))
                for k in sheet.constants:
                    receipts.append(Cell(row_id, k, sheet.name,
                                         f"A{sheet.note_rows[3] if len(sheet.note_rows) > 3 else 1}",
                                         sheet.constants[k], None))
                if entity is not None:
                    receipts.append(Cell(row_id, spec.entity, sheet.name, label_a1,
                                         raw_label, None))
                if sheet.section_dimension and sec is not None:
                    receipts.append(Cell(row_id, sheet.section_dimension, sheet.name,
                                         section_a1, sec, None))
                row_id += 1

    _write(db_path, spec, dims, rows, receipts, notes, {spec.measure})
    return {"rows": len(rows), "receipts": len(receipts), "notes": len(notes),
            "formulas": sum(1 for c in receipts if c.formula)}


def _ingest_record(spec: Spec, wb_v, wb_f, db_path: Path) -> dict:
    """A tidy table stays tidy: one output row per sheet row, columns kept as they are.

    No unpivot, no period, no section, no total-classification — a record table has none of that.
    Each identifier cell and each measure cell records its own address in the same iteration it is
    read (D6), exactly as the crosstab path does. What differs is only the shape.
    """
    from openpyxl.utils import column_index_from_string
    rows: list[dict] = []
    receipts: list[Cell] = []
    notes: list[tuple[str, str, str]] = []
    row_id = 0
    id_names: list[str] = []
    measure_names: list[str] = []

    for sheet in spec.sheets:
        sv, sf = wb_v[sheet.name], wb_f[sheet.name]
        idx = {c: column_index_from_string(c) for c in sheet.id_columns + sheet.measure_columns}

        def header_name(col_letter: str) -> str:
            raw = sv.cell(sheet.header_row, idx[col_letter]).value
            return clean_label(str(raw)) if raw is not None else col_letter

        sheet_ids = [header_name(c) for c in sheet.id_columns]
        sheet_meas = [header_name(c) for c in sheet.measure_columns]
        # First sheet fixes the column order; a second record sheet must line up (kept simple).
        if not id_names:
            id_names, measure_names = sheet_ids, sheet_meas

        for r in range(sheet.first_data_row, sheet.last_data_row + 1):
            # Skip a fully-blank row; a record sheet has no section headers to worry about.
            if all(sv.cell(r, idx[c]).value in (None, "") for c in sheet.id_columns):
                continue
            rec: dict = {"__row_id": row_id, "row_kind": "entity"}
            for name, letter in zip(sheet_ids, sheet.id_columns):
                v = sv.cell(r, idx[letter]).value
                rec[name] = None if v is None else str(v).strip()
                if v is not None:
                    receipts.append(Cell(row_id, name, sheet.name, f"{letter}{r}", str(v), None))
            for name, letter in zip(sheet_meas, sheet.measure_columns):
                v = sv.cell(r, idx[letter]).value
                rec[name] = float(v) if isinstance(v, (int, float)) else None
                if isinstance(v, (int, float)):
                    f = sf.cell(r, idx[letter]).value
                    receipts.append(Cell(row_id, name, sheet.name, f"{letter}{r}", str(v),
                                         f if isinstance(f, str) and f.startswith("=") else None))
            rows.append(rec)
            row_id += 1

    dims = id_names + measure_names + ["row_kind"]
    _write(db_path, spec, dims, rows, receipts, notes, set(measure_names))
    return {"rows": len(rows), "receipts": len(receipts), "notes": len(notes),
            "formulas": sum(1 for c in receipts if c.formula)}


def _write(db_path: Path, spec: Spec, dims: list[str], rows, receipts, notes,
           numeric: set) -> None:
    con = duckdb.connect(str(db_path))
    # Column names reach SQL as double-quoted identifiers, and an embedded `"` is the one character
    # that breaks out of the quotes — e.g. a spec `entity` or a crafted header cell of
    # `x" DOUBLE); DROP TABLE cell_map; --`. Real labels (even `% vs pivot (now)`) never contain a
    # `"`, so refuse it here, the one point where every column name — spec-derived and
    # workbook-derived — converges before the CREATE.
    bad = [d for d in dims if '"' in d]
    if bad:
        raise ValueError(f"column name(s) {bad} contain a double-quote and cannot be a safe "
                         f"SQL identifier")

    # Double-count guard (D97). If two rows share the same grain — every dimension except the
    # measure — there is nothing to tell them apart, so a later "sum by entity" silently adds them
    # together. That happens when variant sheets (product MS / HSD / ALL) were ingested with no
    # distinguishing constant. Refuse now with a concrete hint rather than answer a wrong number.
    grain = [d for d in dims if d not in numeric]
    if grain and rows:
        from collections import Counter
        dup = sum(1 for c in Counter(tuple(r.get(d) for d in grain) for r in rows).values() if c > 1)
        if dup:
            raise ValueError(
                f"{dup} rows share the same ({', '.join(grain)}) with nothing to tell them apart — "
                f"these sheets look like the same table repeated. On the confirm screen, set what "
                f"distinguishes them (e.g. product=MS on one sheet, product=HSD on another), or the "
                f"numbers would be double-counted.")

    cols = ", ".join(f'"{d}" {"DOUBLE" if d in numeric else "VARCHAR"}' for d in dims)
    for t in (spec.table, "cell_map", "sheet_notes"):
        con.execute(f"DROP TABLE IF EXISTS {t}")
    con.execute(f"CREATE TABLE {spec.table} (__row_id INTEGER, {cols})")
    con.execute("""CREATE TABLE cell_map (table_name VARCHAR, row_id INTEGER,
                   column_name VARCHAR, sheet VARCHAR, a1 VARCHAR,
                   raw_value VARCHAR, formula VARCHAR)""")
    con.execute("CREATE TABLE sheet_notes (sheet VARCHAR, a1 VARCHAR, text VARCHAR)")

    marks = ",".join("?" * (len(dims) + 1))
    # DuckDB's executemany rejects an empty parameter list, so guard each insert. A workbook with
    # no footnote/source rows (a plain export, or a synthetic test file) has empty `notes` and
    # would otherwise crash ingest with "requires a non-empty list".
    if rows:
        con.executemany(f"INSERT INTO {spec.table} VALUES ({marks})",
                        [tuple([r["__row_id"]] + [r.get(d) for d in dims]) for r in rows])
    if receipts:
        con.executemany(f"INSERT INTO cell_map VALUES ('{spec.table}',?,?,?,?,?,?)",
                        [(c.row_id, c.column, c.sheet, c.a1, c.raw_value, c.formula)
                         for c in receipts])
    if notes:
        con.executemany("INSERT INTO sheet_notes VALUES (?,?,?)", notes)
    con.close()


if __name__ == "__main__":
    root = Path(__file__).resolve().parent.parent
    spec = Spec.load(Path(sys.argv[1]) if len(sys.argv) > 1 else root / "specs" / "ppac.yaml")
    out = ingest(spec, root / "data" / spec.file,
                 root / "data" / f"{Path(spec.file).stem}.duckdb")
    print(f"{spec.file}: " + " · ".join(f"{v} {k}" for k, v in out.items()))
