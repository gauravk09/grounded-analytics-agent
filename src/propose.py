"""Propose a spec for a workbook nobody has seen. Heuristics first, model second, human last.

The division of labour is the whole design:

    heuristics  GEOMETRY — where the header is, which columns hold values, which rows are
                totals. Derivable from structure, and cheap. Where these are confident, NO
                MODEL RUNS.

    model       MEANING — what the label column is ("state", "account head"), what to call the
                unpivoted header dimension ("year"), aliases, descriptions. Not derivable from
                structure; it is language.

    human       confirms once. Then the spec is frozen and ingest is deterministic forever.

The model sees the STRUCTURE SKETCH, never the values — the same rule the catalog follows (D17).
It is looking at a picture of the sheet's shape plus its labels, which is a perception task with
a small checkable answer, not arithmetic.
"""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl.utils import get_column_letter

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent))
from probe import TOTAL_WORDS, SheetProfile, profile, sheets         # noqa: E402
from catalog import PERIOD_PATTERNS                                 # noqa: E402
from spec import SheetSpec, Spec                                     # noqa: E402

HIGH, MEDIUM, LOW = "high", "medium", "low"


# ── geometry: mechanical, no model ──────────────────────────────────────────────
def value_columns(p: SheetProfile) -> tuple[int, int] | None:
    """Columns that predominantly hold numbers across the sheet."""
    cols = sorted(c for c, n in p.col_numeric.items() if n >= 3 and n > p.col_text.get(c, 0))
    if not cols:
        # A very short table — a single time series has one value per period column, so the
        # depth>=3 rule finds nothing. Fall back to a CONTIGUOUS run of >=3 numeric columns; the
        # contiguity requirement keeps a stray number in a title row from looking like a value block.
        weak = sorted(c for c, n in p.col_numeric.items() if n >= 1 and n >= p.col_text.get(c, 0))
        if len(weak) >= 3 and weak[-1] - weak[0] == len(weak) - 1:
            cols = weak
    return (cols[0], cols[-1]) if cols else None


def data_rows(p: SheetProfile, lo: int, hi: int) -> tuple[int, int] | None:
    """Rows where at least half the value columns actually hold numbers.

    Deliberately not "the row is mostly numeric" — the Budget sheet has two label columns and
    stray text markers ('..') inside numeric columns, so a whole-row test misreads it.
    """
    width = hi - lo + 1
    hits = [r.row for r in p.rows
            if sum(1 for ch in (r.cells[lo - 1:hi] if r.cells else "") if ch in "#f")
            >= max(1, width // 2)]
    return (hits[0], hits[-1]) if hits else None


def header_row(p: SheetProfile, lo: int, hi: int, first_data: int) -> int | None:
    """The last row above the data whose VALUE columns hold text.

    Not "the last mostly-text row": in PPAC the header is text all the way across, but in the
    Budget sheet only columns D-G carry it. What both share is that the header names the value
    columns, so that is what to look for.
    """
    for r in reversed([r for r in p.rows if r.row < first_data]):
        if any(ch == "T" for ch in (r.cells[lo - 1:hi] if r.cells else "")):
            return r.row
    return None


def label_columns(p: SheetProfile, first_data: int, last_data: int,
                  value_lo: int) -> list[int]:
    """The run of text columns immediately left of the values.

    Label columns come in runs, and the one NEAREST the numbers is the most specific. The Union
    Budget carries the Hindi name in D and the English in E; taking the leftmost read every
    label in Hindi, so nothing matched. Taking the rightmost is not an English preference — it
    is "the column closest to the number it labels", which holds for numbering-then-name and
    category-then-item too.
    """
    counts: dict[int, int] = {}
    for r in p.rows:
        if first_data <= r.row <= last_data and r.cells:
            for i, ch in enumerate(r.cells, start=1):
                if ch == "T":
                    counts[i] = counts.get(i, 0) + 1
    span = last_data - first_data + 1
    good = [c for c in sorted(counts) if counts[c] >= span * 0.5 and c < value_lo]
    if not good:
        return []
    run = [good[-1]]                       # walk left while columns stay contiguous
    while run[0] - 1 in good:
        run.insert(0, run[0] - 1)
    return run


def classify_rows(p: SheetProfile, first_data: int, last_data: int,
                  label_col: int = 1) -> dict[str, str]:
    """Patterns for rows that are aggregates or section headings rather than plain entities."""
    kinds: dict[str, str] = {}

    def lbl(r) -> str | None:
        return (r.texts or {}).get(label_col) or r.label

    totals = [lbl(r) for r in p.rows
              if first_data <= r.row <= last_data and lbl(r) and TOTAL_WORDS.search(lbl(r))]
    if totals:
        repeated = {t for t in totals if totals.count(t) > 1}
        grand = [t for t in totals if t not in repeated]
        if repeated:
            kinds["subtotal"] = "^" + re.escape(sorted(repeated)[0]) + "$"
        if grand:
            # Several distinct one-off totals is normal — a budget sheet has many. One
            # alternation covers them without inventing a pattern per label.
            kinds["grand_total"] = "^(" + "|".join(re.escape(g) for g in sorted(grand)) + ")$"
    sections = [r for r in p.rows
                if first_data <= r.row <= last_data and r.numeric == 0 and lbl(r)
                and not TOTAL_WORDS.search(lbl(r))]
    if len(sections) >= 2:
        prefix = _common_prefix([lbl(s) for s in sections])
        if len(prefix) >= 3:
            kinds["section_header"] = "^" + re.escape(prefix)
    return kinds


def _common_prefix(labels: list[str]) -> str:
    if not labels:
        return ""
    out = labels[0]
    for l in labels[1:]:
        while not l.upper().startswith(out.upper()) and out:
            out = out[:-1]
    return out


def note_rows(p: SheetProfile, first_data: int, last_data: int) -> list[int]:
    """Sparse rows outside the data block — titles, units, sources, footnotes. Citable (D11)."""
    return [r.row for r in p.rows
            if (r.row < first_data or r.row > last_data) and 0 < r.filled <= 2 and r.label]


def propose_geometry(path, sheet: str) -> tuple[SheetSpec | None, dict[str, str], list[str]]:
    p = profile(path, sheet)
    conf: dict[str, str] = {}
    notes: list[str] = []

    vc = value_columns(p)
    if not vc:
        return None, {"value_columns": LOW}, [f"{sheet}: found no numeric columns"]
    lo, hi = vc
    conf["value_columns"] = HIGH if hi - lo >= 2 else MEDIUM

    dr = data_rows(p, lo, hi)
    if not dr:
        return None, {"data_rows": LOW}, [f"{sheet}: found no rows of numbers"]
    first_data, last_data = dr
    # Walk back over label-only rows immediately above the first numeric row: those are
    # section headings, and they belong INSIDE the block. Without this, PPAC's first section
    # (row 9) falls outside while the other four sit inside — the same kind of row classified
    # two different ways depending on where it happens to be.
    by_row = {r.row: r for r in p.rows}

    def is_section(r) -> bool:
        """A label-only row INSIDE the table. Stop at the header — it also has no numbers, but
        it has text in the VALUE columns, which is exactly what distinguishes the two."""
        if r is None or r.numeric or not r.label or r.merged:
            return False
        return not any(ch == "T" for ch in (r.cells[lo - 1:hi] if r.cells else ""))

    while is_section(by_row.get(first_data - 1)):
        first_data -= 1
    conf["data_rows"] = HIGH

    hdr = header_row(p, lo, hi, first_data)
    conf["header_row"] = HIGH if hdr and hdr == first_data - 1 else MEDIUM
    if hdr and hdr < first_data - 1:
        notes.append(f"{sheet}: header on row {hdr} but data starts at {first_data} — "
                     f"check for a second header row or spacer rows between them")

    labs = label_columns(p, first_data, last_data, lo)
    lab = labs[-1] if labs else None
    conf["label_column"] = HIGH if lab else LOW
    if len(labs) > 1:
        others = ", ".join(get_column_letter(c) for c in labs[:-1])
        notes.append(f"{sheet}: using column {get_column_letter(lab)} for names; {others} also "
                     f"holds text (numbering, or the same name in another language)")
    elif lab and lab > 1:
        notes.append(f"{sheet}: labels are in column {get_column_letter(lab)}, not A")

    # Walk up from the header row while the value columns still hold text: a stacked header
    # (the Union Budget's year row sits above its estimate-type row). One row -> leave empty.
    hrun = []
    rr = hdr or first_data - 1
    by_row_all = {r.row: r for r in p.rows}
    while rr >= 1:
        row = by_row_all.get(rr)
        if row and any(ch == "T" for ch in (row.cells[lo - 1:hi] if row.cells else "")):
            hrun.insert(0, rr); rr -= 1
        else:
            break
    header_rows = hrun if len(hrun) > 1 else []

    layout, layout_conf, layout_why = classify_layout(p, hdr or first_data - 1, lo, hi, labs)
    conf["layout"] = layout_conf
    notes.append(f"{sheet}: read as a {layout} — {layout_why}")

    ss = SheetSpec(
        name=sheet,
        layout=layout,
        header_row=hdr or first_data - 1,
        header_rows=header_rows,
        label_column=get_column_letter(lab or 1),
        alt_label_columns=[get_column_letter(c) for c in labs[:-1]],
        first_data_row=first_data,
        last_data_row=last_data,
        first_value_column=get_column_letter(lo),
        last_value_column=get_column_letter(hi),
        note_rows=note_rows(p, first_data, last_data),
        row_kinds=classify_rows(p, first_data, last_data, lab or 1),
    )
    if layout == "record":
        # every text column in the run identifies the row; every numeric column is a measure.
        # No unpivot, no total-classification — a record table has none of that structure.
        ss.id_columns = [get_column_letter(c) for c in labs]
        ss.measure_columns = [get_column_letter(c) for c in range(lo, hi + 1)]
        ss.row_kinds = {}
    return ss, conf, notes



def classify_layout(p: SheetProfile, hdr: int, lo: int, hi: int,
                    label_run: list[int]) -> tuple[str, str, str]:
    """crosstab vs record — the block-2 test, mechanical.

    Look at the header cells sitting over the VALUE columns. If they are values of one variable
    (years: 2008-09, 2009-10 ...) the sheet is a folded cross-tab and ingest must unpivot. If they
    are names of different things (% vs pivot, Price vs 50-DMA) the sheet is already a tidy record
    and ingest must leave it alone.

    Confidence follows AGREEMENT, not existence (the calibration bug): a header run that is cleanly
    all-periodic or cleanly all-text is `high`; a mixed header (2019, 2020, Change %) is `low`, which
    promotes the field to a question on the confirmation screen instead of a silent guess.
    """
    hdr_row = next((r for r in p.rows if r.row == hdr), None)
    labels = [(hdr_row.texts or {}).get(c) for c in range(lo, hi + 1)] if hdr_row else []
    labels = [l for l in labels if l]
    if not labels:
        return "crosstab", LOW, "no header labels over the value columns to judge"

    def is_periodic(lbl: str) -> bool:
        return any(re.match(pat, str(lbl).strip()) for pat in PERIOD_PATTERNS)

    periodic = sum(is_periodic(l) for l in labels)
    frac = periodic / len(labels)
    n_ids = len(label_run)
    # A "clean" header is single-line latin text. Merged, multi-line or non-latin headers are the
    # hard melt-then-split case (the Union Budget keeps the year in a row above and the estimate
    # type — in Hindi and English — in the header row). Those are NOT confident records: a wrong
    # flip to record would stop the file unpivoting and break a workbook that currently works. So
    # a record verdict requires clean headers; anything messier stays crosstab and drops to LOW,
    # which turns it into a question on the confirmation screen instead of a silent guess.
    clean = all("\n" not in l and sum(c.isascii() for c in l) >= 0.8 * len(l) for l in labels)

    if frac >= 0.8:
        conf = HIGH if frac == 1.0 and (hi - lo) >= 2 else MEDIUM
        return "crosstab", conf, f"value headers look like periods ({periodic}/{len(labels)})"
    if frac == 0.0 and clean and n_ids >= 1:
        # Several heterogeneous headers is a confident record; a SINGLE non-periodic value column
        # (a two-column list like City|Population) is still a record — there is just nothing to
        # cross-check, so it is medium confidence and gets confirmed rather than asserted.
        conf = HIGH if len(labels) >= 2 else MEDIUM
        return "record", conf, (f"value header(s) are names, not periods; "
                                f"{n_ids} identifier column(s) to the left")
    if not clean:
        return "crosstab", LOW, ("value headers are merged/multi-line — cannot tell periods from "
                                 "names; defaulting to unpivot, flag for confirmation")
    return "crosstab", LOW, (f"mixed header — {periodic}/{len(labels)} look periodic; "
                             "could be periods plus a derived column, flag for confirmation")


# ── meaning: the model's job ────────────────────────────────────────────────────
MEANING_PROMPT = """You are naming the dimensions of a spreadsheet so an analyst can ask
questions about it. You are shown its STRUCTURE, never its numbers.

In the sketch: T = a text cell, # = a number, f = a formula, . = empty. The text after each row
is that row's leftmost label.

Return JSON only:
{
  "entity":  "what one row is (a single lowercase word: state, product, customer, account)",
  "period":  "what the header row's columns are (year, quarter, month, period)",
  "measure": "what the numbers are (a single lowercase word: value, sales, amount, count)",
  "unit":    "the unit if the sheet states one, else null",
  "annotations": {"<column>": {"description": "one clause", "aliases": ["other words a user
                   might use"]}},
  "absent_concepts": ["concepts a user will plausibly ask about that this sheet does NOT
                      contain — e.g. price, revenue, population"],
  "section_dimension": "if section-header rows group the rows beneath them, what that grouping
                        is called (region, category), else null"
}
Use the sheet's own vocabulary. Do not invent columns."""


def propose_meaning(sketch: str, planner=None) -> dict:
    """Ask a model to name the dimensions. Returns {} if no model is available.

    Failing to name things is survivable — the spec still ingests correctly, the assistant is
    just worse at mapping a user's words. Failing at GEOMETRY is not survivable, which is why
    that half never depends on a model.
    """
    if planner is None:
        return {}
    try:
        from planner import DeepSeekPlanner
        tier = next((t for t in getattr(planner, "tiers", [planner])
                     if isinstance(t, DeepSeekPlanner)), None)
        if tier is None:
            return {}
        raw = tier.client.chat.completions.create(
            model=tier.model,
            messages=[{"role": "system", "content": MEANING_PROMPT},
                      {"role": "user", "content": sketch}],
            response_format={"type": "json_object"}, temperature=0,
            extra_body=tier.extra,
        ).choices[0].message.content
        return json.loads(raw)
    except Exception:
        return {}


def verify_absent(concepts: list[str], path, specs) -> tuple[list[str], list[str]]:
    """Drop any 'absent' concept the sheet demonstrably contains.

    `absent_concepts` causes a HARD refusal before any model runs, so a wrong entry silently
    kills real questions. On the Union Budget the model proposed
    ['revenue', 'expenditure', 'deficit'] — the file's three main subjects.

    The check is mechanical: a word that appears in the sheet's own labels is not absent. This
    is the same principle as everywhere else — the model may propose, but a claim that can be
    checked against the data gets checked.
    """
    vocabulary = set()
    for sp in specs:
        prof = profile(path, sp.name)
        for r in prof.rows:
            for t in (r.texts or {}).values():
                vocabulary.update(re.findall(r"[a-z]{3,}", t.lower()))
    kept, rejected = [], []
    for c in concepts:
        words = re.findall(r"[a-z]{3,}", c.lower())
        (rejected if words and all(w in vocabulary for w in words) else kept).append(c)
    return kept, rejected


def propose(path, planner=None, table: str | None = None) -> Spec:
    """Full proposal: geometry from structure, meaning from a model, both marked for review."""
    path = Path(path)
    specs, conf, notes = [], {}, []
    for sheet in sheets(path):
        s, c, n = propose_geometry(path, sheet)
        notes += n
        for k, v in c.items():
            conf[f"{sheet}.{k}"] = v
        if s:
            specs.append(s)
        else:
            notes.append(f"{sheet}: skipped — no usable table found")

    if not specs:
        raise ValueError(f"no sheet in {path.name} looks like a table")

    # Anchors, not the first N rows. Truncation hid exactly the rows that matter — grand totals
    # and source footnotes sit at the BOTTOM of every sheet. And because cost now tracks how
    # irregular a sheet is rather than how long it is, every sheet fits; the old `[:3]` cap meant
    # the model never saw sheets 4 and 5 of the Budget at all.
    sketch = f"File: {path.name}\n\n" + "\n\n".join(
        f"Sheet {s.name!r}:\n{profile(path, s.name).sketch(cols=12)}" for s in specs)
    meaning = propose_meaning(sketch, planner)

    for s in specs:
        s.section_dimension = meaning.get("section_dimension") or None

    # A record workbook's identifier and measures are real columns; entity/period/measure must name
    # them or every few-shot example teaches the model a schema it does not have. Read the header
    # names for the first sheet's id/measure columns.
    if any(s.layout == "record" for s in specs):
        from openpyxl.utils import column_index_from_string
        from xlsx_io import load as _load
        s0 = next(s for s in specs if s.layout == "record")
        wsv = _load(path, data_only=True)[s0.name]
        def _hname(letter):
            v = wsv.cell(s0.header_row, column_index_from_string(letter)).value
            return str(v).strip() if v is not None else letter
        ent_name = _hname(s0.id_columns[0]) if s0.id_columns else "entity"
        meas_name = _hname(s0.measure_columns[0]) if s0.measure_columns else "value"
    else:
        ent_name, meas_name = None, None

    absent, rejected = verify_absent(meaning.get("absent_concepts") or [], path, specs)
    if rejected:
        notes.append(f"dropped {rejected} from absent_concepts — those words appear in the "
                     f"sheet's own labels, so the file plainly does contain them")

    # Guarantee a plain SQL identifier so a numeric-leading filename ("2024.xlsx" -> "2024")
    # doesn't trip the Spec.table validator; underscore prefix keeps it a legal, unquoted name.
    stem = re.sub(r"\W+", "_", path.stem).lower()
    if not re.match(r"[A-Za-z_]", stem):
        stem = "_" + stem
    return Spec(
        file=path.name,
        table=table or stem,
        entity=meaning.get("entity") or ent_name or "entity",
        period=meaning.get("period") or "period",
        measure=meaning.get("measure") or meas_name or "value",
        unit=meaning.get("unit"),
        sheets=specs,
        annotations=meaning.get("annotations") or {},
        absent_concepts=absent,
        confidence=conf,
        notes_for_reviewer=notes,
    )


if __name__ == "__main__":
    from planner import get_planner
    p = Path(sys.argv[1])
    use_model = "--no-model" not in sys.argv
    spec = propose(p, get_planner() if use_model else None)
    print(spec.model_dump_json(indent=2, exclude_none=True))
